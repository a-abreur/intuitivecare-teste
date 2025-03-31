import os
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path
import zipfile
import shutil

# ====================== CONFIGURAÇÃO ======================
INPUT_PDF = Path("../web_scraping/output/Anexo_I.pdf")  # Altere para seu PDF
OUTPUT_DIR = Path("output_tabelas")
OUTPUT_XLSX_TABELAS = OUTPUT_DIR / "Todas_Tabelas_Formatadas.xlsx"
OUTPUT_XLSX_CONSOLIDADO = OUTPUT_DIR / "Dados_Consolidados.xlsx"
ZIP_FILE = "Teste_Arthur.zip"

# ====================== FUNÇÕES ======================
def extract_all_tables(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        return [table for page in pdf.pages for table in page.extract_tables() if table and len(table) > 1]

def apply_excel_formatting(ws):
    styles = {
        'header': {'font': Font(bold=True, color="FFFFFF", size=11), 'fill': PatternFill(start_color="00B4B4", fill_type="solid")},
        'data': {'font': Font(size=10), 'fill_gray': PatternFill(start_color="D3D3D3", fill_type="solid"), 'fill_white': PatternFill(start_color="FFFFFF", fill_type="solid")},
        'align': {'center': Alignment(horizontal="center", vertical="center", wrap_text=True), 'left': Alignment(horizontal="left", vertical="center", wrap_text=True)},
        'border': Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    }
    
    for cell in ws[1]:
        cell.font, cell.fill, cell.alignment, cell.border = styles['header']['font'], styles['header']['fill'], styles['align']['center'], styles['border']
    
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font, cell.alignment, cell.border = styles['data']['font'], styles['align']['left'], styles['border']
            cell.fill = styles['data']['fill_gray'] if cell.row % 2 == 0 else styles['data']['fill_white']
    
    adjust_column_widths(ws)

def adjust_column_widths(ws):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(cell.value)) if cell.value else 10) for cell in col)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

def save_excel(tables, output_path, consolidated=False):
    wb = Workbook()
    wb.remove(wb.active) if not consolidated else None
    
    if consolidated:
        ws = wb.create_sheet(title="Dados Consolidados")
        ws.append(tables[0][0])
        for table in tables:
            for row in table[1:]:
                ws.append(row)
        apply_excel_formatting(ws)
    else:
        for i, table in enumerate(tables, 1):
            ws = wb.create_sheet(title=f"Tabela_{i}")
            for row in table:
                ws.append(row)
            apply_excel_formatting(ws)
    
    wb.save(output_path)

def create_zip_with_files(zip_filename, files_to_zip):
    with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for file in files_to_zip:
            if Path(file).exists():
                zipf.write(file, arcname=Path(file).name)
                print(f"📦 Arquivo {file} adicionado ao ZIP")
            else:
                print(f"⚠️ Arquivo não encontrado: {file}")

# ====================== EXECUÇÃO PRINCIPAL ======================
if __name__ == "__main__":
    try:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        if not INPUT_PDF.exists():
            raise FileNotFoundError(f"Arquivo PDF não encontrado: {INPUT_PDF}")
        
        print("⏳ Processando PDF...")
        todas_tabelas = extract_all_tables(INPUT_PDF)
        if not todas_tabelas:
            print("❌ Nenhuma tabela encontrada!")
        else:
            print(f"📊 {len(todas_tabelas)} tabelas encontradas")
            
            save_excel(todas_tabelas, OUTPUT_XLSX_TABELAS)
            save_excel(todas_tabelas, OUTPUT_XLSX_CONSOLIDADO, consolidated=True)
            
            files_to_zip = [OUTPUT_XLSX_TABELAS, OUTPUT_XLSX_CONSOLIDADO]
            create_zip_with_files(ZIP_FILE, files_to_zip)
            
            for file in files_to_zip:
                os.remove(file)
            
            print(f"🎉 Processo concluído com sucesso! Arquivo ZIP criado: {ZIP_FILE}")
    except Exception as e:
        print(f"❌ Erro: {str(e)}")